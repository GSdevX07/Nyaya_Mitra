"""
seed_supabase.py
────────────────
Reads Nyaya_Mitra_Synthetic_Dataset.xlsx and:
  1. Creates all required tables in Supabase (via REST API using service key)
  2. Seeds all rows from every sheet into Supabase tables

Tables created / seeded:
  - offenses        (lookup)
  - jails           (lookup)
  - lawyers_lookup  (lookup)
  - undertrial_cases
  - documents
  - bail_applications
  - status_tracking
"""

import os, sys, json
from datetime import datetime, date
import openpyxl
from supabase import create_client, Client
from dotenv import load_dotenv

# ── Load env ──────────────────────────────────────────────────────────────────
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
EXCEL_PATH   = os.path.join(os.path.dirname(__file__), "..", "Nyaya_Mitra_Synthetic_Dataset.xlsx")

sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Helpers ───────────────────────────────────────────────────────────────────
def read_sheet(wb, sheet_name, header_row=3):
    """Read a sheet where headers are on header_row (1-indexed)."""
    ws = wb[sheet_name]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            if h is None:
                continue
            v = ws.cell(r, c).value
            # Convert date objects to ISO string
            if isinstance(v, (datetime, date)):
                v = v.isoformat()
            row[h] = v
        # Skip completely empty rows
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows

def upsert_batch(table: str, rows: list, on_conflict: str, chunk=50):
    """Upsert rows in chunks to avoid payload limits."""
    total = len(rows)
    for i in range(0, total, chunk):
        batch = rows[i:i+chunk]
        res = sb.table(table).upsert(batch, on_conflict=on_conflict).execute()
        print(f"  [{table}] upserted rows {i+1}–{min(i+chunk, total)}/{total}")
    print(f"  ✓ {table}: {total} rows seeded")

# ── Load Excel ────────────────────────────────────────────────────────────────
print("Loading Excel dataset …")
wb = openpyxl.load_workbook(EXCEL_PATH)

# ── 1. Offenses (lookup) ──────────────────────────────────────────────────────
print("\n[1/7] Seeding offenses …")
offense_rows = read_sheet(wb, "Offenses (lookup)")
# Map: offense_code, section, description, max_sentence_days
offenses = [
    {
        "offense_code": r["offense_code"],
        "section":      r["section"],
        "description":  r["description"],
        "max_sentence_days": int(r["max_sentence_days"]),
    }
    for r in offense_rows
]
upsert_batch("offenses", offenses, "offense_code")

# ── 2. Jails (lookup) ─────────────────────────────────────────────────────────
print("\n[2/7] Seeding jails …")
jail_rows = read_sheet(wb, "Jails (lookup)")
jails = [
    {
        "jail_id":       r["jail_id"],
        "jail_name":     r["jail_name"],
        "state":         r["state"],
        "occupancy_pct": int(r["occupancy_pct"]) if r["occupancy_pct"] is not None else None,
    }
    for r in jail_rows
]
upsert_batch("jails", jails, "jail_id")

# ── 3. Lawyers lookup ─────────────────────────────────────────────────────────
print("\n[3/7] Seeding lawyers_lookup …")
lawyer_rows = read_sheet(wb, "Lawyers (lookup)")
lawyers_lookup = [
    {
        "lawyer_id":     r["lawyer_id"],
        "lawyer_name":   r["lawyer_name"],
        "dlsa_district": r["dlsa_district"],
    }
    for r in lawyer_rows
]
upsert_batch("lawyers_lookup", lawyers_lookup, "lawyer_id")

# ── 4. Undertrial Cases ───────────────────────────────────────────────────────
print("\n[4/7] Seeding undertrial_cases …")
case_rows = read_sheet(wb, "Cases (200 rows)")
cases = []
for r in case_rows:
    # Split present_docs CSV string into array
    pd = r.get("present_docs") or ""
    present_docs = [d.strip() for d in pd.split(",") if d.strip()] if pd else []

    cases.append({
        "id":                          r["case_id"],
        "offense_code":                r["offense_code"],
        "jail_id":                     r["jail_id"],
        "lawyer_id":                   r.get("lawyer_id"),
        "arrest_date":                 r["arrest_date"],
        "custody_days":                int(r["custody_days"]),
        "max_sentence_days_for_offense": int(r["max_sentence_days"]),
        "eligibility_threshold_days":  int(r["eligibility_threshold_days"]),
        "days_overdue":                int(r["days_overdue"]),
        "eligibility_status":          r["eligibility_status"],
        "first_time_offender":         bool(r["first_time_offender"]),
        "age":                         int(r["age"]),
        "health_flag":                 bool(r["health_flag"]),
        "preferred_language":          r["preferred_language"],
        "present_docs":                present_docs,
        "records_complete":            bool(r["records_complete"]),
        "urgency_score":               float(r["urgency_score"]),
        "name":                        "synthetic - not a real person",
        "status":                      "DISCOVERED",
        "assignment_status":           "AVAILABLE",
    })
upsert_batch("undertrial_cases", cases, "id")

# ── 5. Documents ──────────────────────────────────────────────────────────────
print("\n[5/7] Seeding documents …")
doc_rows = read_sheet(wb, "Documents")
documents = [
    {
        "id":            r["doc_id"],
        "case_id":       r["case_id"],
        "document_type": r["doc_type"],
        "status":        r["status"],
        "is_present":    r["status"] == "Present",
    }
    for r in doc_rows
]
upsert_batch("documents", documents, "id")

# ── 6. Bail Applications ──────────────────────────────────────────────────────
print("\n[6/7] Seeding bail_applications …")
bail_rows = read_sheet(wb, "Bail_Applications")
bail_apps = [
    {
        "id":                        r["application_id"],
        "case_id":                   r["case_id"],
        "filed_date":                r["filed_date"],
        "status":                    r["status"],
        "next_hearing_or_order_date": r.get("next_hearing_or_order_date"),
    }
    for r in bail_rows
]
upsert_batch("bail_applications", bail_apps, "id")

# ── 7. Status Tracking ────────────────────────────────────────────────────────
print("\n[7/7] Seeding status_tracking …")
tracking_rows = read_sheet(wb, "Status_Tracking")
tracking = [
    {
        "id":             r["tracking_id"],
        "application_id": r["application_id"],
        "event":          r["event"],
        "event_date":     r["event_date"],
    }
    for r in tracking_rows
]
upsert_batch("status_tracking", tracking, "id")

print("\n\n✅  All tables seeded successfully into Supabase!")
print(f"   Project: {SUPABASE_URL}")
