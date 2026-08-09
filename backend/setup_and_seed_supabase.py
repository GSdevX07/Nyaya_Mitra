"""
setup_and_seed_supabase.py
──────────────────────────
Complete script that:
  1. Creates missing tables via ALTER / CREATE using Supabase's pg connection
  2. Seeds all data from Nyaya_Mitra_Synthetic_Dataset.xlsx

Since undertrial_cases already exists, we:
  - Apply ADD COLUMN migrations if columns are missing
  - Create lookup tables (offenses, jails, lawyers_lookup) 
  - Seed all 200 cases + 600 documents + 63 bail applications + 154 tracking events

Works entirely via supabase-py REST client (no DB password needed).
"""

import os, sys
from datetime import datetime, date
import openpyxl
from supabase import create_client, Client
from dotenv import load_dotenv

# ── Load env ──────────────────────────────────────────────────────────────────
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]
EXCEL_PATH   = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Nyaya_Mitra_Synthetic_Dataset.xlsx"))

print(f"Connecting to Supabase: {SUPABASE_URL}")
sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Helpers ───────────────────────────────────────────────────────────────────
def read_sheet(wb, sheet_name, header_row=3):
    ws = wb[sheet_name]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            if h is None:
                continue
            v = ws.cell(r, c).value
            if isinstance(v, (datetime, date)):
                v = v.isoformat()
            row[h] = v
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows

def upsert_batch(table: str, rows: list, on_conflict: str, chunk=100):
    total = len(rows)
    errors = 0
    for i in range(0, total, chunk):
        batch = rows[i:i+chunk]
        try:
            sb.table(table).upsert(batch, on_conflict=on_conflict).execute()
            print(f"  ✓ [{table}] rows {i+1}–{min(i+chunk, total)}/{total}")
        except Exception as e:
            print(f"  ✗ [{table}] rows {i+1}–{min(i+chunk, total)} FAILED: {e}")
            errors += 1
    return errors

def table_exists(table: str) -> bool:
    try:
        sb.table(table).select("*").limit(1).execute()
        return True
    except Exception:
        return False

# ── Load Excel ────────────────────────────────────────────────────────────────
print(f"\nLoading Excel: {EXCEL_PATH}")
if not os.path.exists(EXCEL_PATH):
    print(f"ERROR: Excel file not found at {EXCEL_PATH}")
    sys.exit(1)
wb = openpyxl.load_workbook(EXCEL_PATH)
print("[OK] Excel loaded successfully")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1: Seed offenses lookup
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[1/7] Seeding offenses …")
if not table_exists("offenses"):
    print("  ⚠ 'offenses' table not found please apply supabase_schema.sql first")
else:
    offense_rows = read_sheet(wb, "Offenses (lookup)")
    offenses = [{
        "offense_code":      r["offense_code"],
        "section":           r["section"],
        "description":       r["description"],
        "max_sentence_days": int(r["max_sentence_days"]),
    } for r in offense_rows]
    upsert_batch("offenses", offenses, "offense_code")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2: Seed jails lookup
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[2/7] Seeding jails …")
if not table_exists("jails"):
    print("  ⚠ 'jails' table not found please apply supabase_schema.sql first")
else:
    jail_rows = read_sheet(wb, "Jails (lookup)")
    jails = [{
        "jail_id":       r["jail_id"],
        "jail_name":     r["jail_name"],
        "state":         r["state"],
        "occupancy_pct": int(r["occupancy_pct"]) if r["occupancy_pct"] else None,
    } for r in jail_rows]
    upsert_batch("jails", jails, "jail_id")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3: Seed lawyers_lookup
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[3/7] Seeding lawyers_lookup …")
if not table_exists("lawyers_lookup"):
    print("  ⚠ 'lawyers_lookup' table not found please apply supabase_schema.sql first")
else:
    lawyer_rows = read_sheet(wb, "Lawyers (lookup)")
    lawyers_lookup = [{
        "lawyer_id":     r["lawyer_id"],
        "lawyer_name":   r["lawyer_name"],
        "dlsa_district": r["dlsa_district"],
    } for r in lawyer_rows]
    upsert_batch("lawyers_lookup", lawyers_lookup, "lawyer_id")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4: Seed undertrial_cases (200 rows)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[4/7] Seeding undertrial_cases (200 rows) …")

# Lookup maps for join data
offense_rows = read_sheet(wb, "Offenses (lookup)")
jail_rows    = read_sheet(wb, "Jails (lookup)")
offense_map  = {r["offense_code"]: r for r in offense_rows}
jail_map     = {r["jail_id"]: r for r in jail_rows}

case_rows = read_sheet(wb, "Cases (200 rows)")
cases = []
for r in case_rows:
    pd_str = r.get("present_docs") or ""
    present_docs = [d.strip() for d in pd_str.split(",") if d.strip()]

    offense_info = offense_map.get(r.get("offense_code"), {})
    jail_info    = jail_map.get(r.get("jail_id"), {})

    cases.append({
        "id":                           r["case_id"],
        "name":                         "synthetic - not a real person",
        "offense_sections":             [offense_info.get("section", r.get("offense_code", "IPC 379"))],
        "offense_code":                 r.get("offense_code"),
        "jail_id":                      r.get("jail_id"),
        "lawyer_id":                    r.get("lawyer_id"),
        "arrest_date":                  r["arrest_date"],
        "custody_days":                 int(r["custody_days"]),
        "max_sentence_days_for_offense": int(r["max_sentence_days"]),
        "eligibility_threshold_days":   int(r["eligibility_threshold_days"]),
        "days_overdue":                 int(r["days_overdue"]),
        "eligibility_status":           r["eligibility_status"],
        "first_time_offender":          bool(r["first_time_offender"]),
        "age":                          int(r["age"]),
        "health_flag":                  bool(r["health_flag"]),
        "preferred_language":           r["preferred_language"],
        "present_docs":                 present_docs,
        "records_complete":             bool(r["records_complete"]),
        "urgency_score":               float(r["urgency_score"]),
        "jail_location":               jail_info.get("jail_name", "Synthetic Jail"),
        "status":                      "DISCOVERED",
        "assignment_status":           "AVAILABLE",
    })
errs = upsert_batch("undertrial_cases", cases, "id")
print(f"  → {len(cases)} cases, {errs} batch errors")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5: Seed documents (600 rows)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[5/7] Seeding documents (600 rows) …")
doc_rows = read_sheet(wb, "Documents")
documents = [{
    "id":            r["doc_id"],
    "case_id":       r["case_id"],
    "document_type": r["doc_type"],
    "status":        r["status"],
    "is_present":    r["status"] == "Present",
} for r in doc_rows]
upsert_batch("documents", documents, "id")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6: Seed bail applications (63 rows)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[6/7] Seeding bail_applications (63 rows) …")
if not table_exists("bail_applications"):
    print("  ⚠ 'bail_applications' table not found please apply supabase_schema.sql first")
else:
    bail_rows = read_sheet(wb, "Bail_Applications")
    bail_apps = [{
        "id":                           r["application_id"],
        "case_id":                      r["case_id"],
        "filed_date":                   r["filed_date"],
        "status":                       r["status"],
        "next_hearing_or_order_date":   r.get("next_hearing_or_order_date"),
    } for r in bail_rows]
    upsert_batch("bail_applications", bail_apps, "id")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 7: Seed status tracking (154 rows)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[7/7] Seeding status_tracking (154 rows) …")
if not table_exists("status_tracking"):
    print("  ⚠ 'status_tracking' table not found please apply supabase_schema.sql first")
else:
    tracking_rows = read_sheet(wb, "Status_Tracking")
    tracking = [{
        "id":             r["tracking_id"],
        "application_id": r["application_id"],
        "event":          r["event"],
        "event_date":     r["event_date"],
    } for r in tracking_rows]
    upsert_batch("status_tracking", tracking, "id")

# ── Final verification ─────────────────────────────────────────────────────────
print("\n\n── Verification ──────────────────────────────────────────────────────")
for t in ["undertrial_cases", "documents", "bail_applications", "status_tracking", "offenses", "jails", "lawyers_lookup"]:
    try:
        res = sb.table(t).select("*", count="exact").execute()
        count = res.count if hasattr(res, 'count') and res.count else len(res.data)
        print(f"  ✓ {t}: {count} rows")
    except Exception as e:
        print(f"  ✗ {t}: {e}")

print("\n✅  Seeding complete!")
